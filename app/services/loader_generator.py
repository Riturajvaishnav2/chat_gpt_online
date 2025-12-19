from __future__ import annotations

import json
from datetime import datetime, timezone
import re
import shutil
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from app.models.schemas import ExcelOutputPlan, GenerateLoaderResponse, LoaderOutput
from app.services.openai_client import generate_json_object
from app.services.text_extract import TextExtractionError, extract_text
from app.utils.files import BASE_DIR, OUTPUT_DIR, create_versioned_output_dir, safe_agreement_base_name

SYSTEM_PROMPT = """You are the DISCOUNT IOT PROCESSING ENGINE — MASTER RULESET.
Return ONLY a single JSON object matching the schema below (no markdown, no code fences, no extra text).

Schema:
{
  "agreement_name": string,
  "standards_used": string[],
  "excel_outputs": [
    {
      "direction": "TI" | "TO",
      "client_tadig": string,
      "partner_tadig": string,
      "start_date": "YYYY-MM-DD",
      "end_date": "YYYY-MM-DD",
      "currency": string,
      "sms_mo_rate": number,
      "sms_mt_rate": number,
      "is_discount": boolean,
      "filename": string
    }
  ],
  "mappings": [
    {
      "clause_id": string,
      "clause_text": string,
      "matched_standard": string | null,
      "confidence": number (0..1),
      "loader_fields": object
    }
  ],
  "missing_fields": string[],
  "notes": string
}

FILENAME FORMAT: <ClientTADIG>_<PartnerTADIG>_<TIorTO>_<StartDateYYYYMMDD>_<EndDateYYYYMMDD>_D.xlsx

RULES:
- Follow the DISCOUNT IOT PROCESSING ENGINE instructions to detect direction (TI vs TO), identify client/partner TADIGs, find mapping header, apply country TADIG logic, pick TAX EXCLUSIVE tables, and extract SMS rates.
- For each sheet, generate ONLY the file that matches the detected direction (TI -> TAP-IN / outbound, TO -> TAP-OUT / inbound).
- SMS extraction: exactly one SMS-MO and one SMS-MT per sheet; if SMS-MT missing, set sms_mt_rate=0.
- Excel outputs must include: direction, client_tadig, partner_tadig, start/end dates, currency, sms_mo_rate, sms_mt_rate, is_discount flag (default true), and filename in the required format.
- Use only provided text; do not invent clause text. If data is missing, set confidence <= 0.4 and add to missing_fields.
- Keep clause_id stable and human-readable (1, 1.1, A-3, or C1/C2...).
"""


MAX_CHARS_PER_FILE = 60_000
MAX_TOTAL_CHARS = 220_000
SMS_MO_ROW = 35
SMS_MT_ROW = 53
SMS_RATE_COL = "J"
SMS_DISCOUNT_COL = "I"
SMS_CURRENCY_COL = "H"
SMS_INITIAL_FEE_COL = "L"
SMS_CHARGING_INTERVAL_COL = "K"
FILENAME_PATTERN = re.compile(
    r"^(?P<client>[A-Za-z0-9]+)_(?P<partner>[A-Za-z0-9]+)_(?P<direction>TI|TO)_(?P<start>\d{8})_(?P<end>\d{8})_D\.xlsx$"
)


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _write_sms_row(ws, row_idx: int, *, rate: float, currency: str, is_discount: bool, highlight: PatternFill) -> None:
    ws[f"{SMS_CURRENCY_COL}{row_idx}"] = currency
    ws[f"{SMS_DISCOUNT_COL}{row_idx}"] = 1 if is_discount else 0
    ws[f"{SMS_RATE_COL}{row_idx}"] = rate
    ws[f"{SMS_INITIAL_FEE_COL}{row_idx}"] = ""
    # Highlight rate + charging interval cells to signal updates without altering the interval value.
    for col in (SMS_RATE_COL, SMS_CHARGING_INTERVAL_COL):
        cell = ws[f"{col}{row_idx}"]
        cell.fill = highlight
    # Clear off-peak related cells to avoid leaking stale template values.
    for col in ("P", "Q", "R"):
        ws[f"{col}{row_idx}"] = None


def _generate_excel_from_template(template_path: Path, dest_path: Path, plan: ExcelOutputPlan) -> None:
    match = FILENAME_PATTERN.match(plan.filename)
    if not match:
        raise HTTPException(
            status_code=422,
            detail=f"Invalid output filename '{plan.filename}'. Expected pattern <ClientTADIG>_<PartnerTADIG>_<TIorTO>_<StartDateYYYYMMDD>_<EndDateYYYYMMDD>_D.xlsx",
        )
    if match.group("direction") != plan.direction:
        raise HTTPException(
            status_code=422,
            detail=f"Filename direction {match.group('direction')} does not match plan direction {plan.direction}",
        )

    start_digits = plan.start_date.replace("-", "")
    end_digits = plan.end_date.replace("-", "")
    if match.group("start") != start_digits or match.group("end") != end_digits:
        raise HTTPException(
            status_code=422,
            detail="Filename start/end dates must match plan start_date/end_date.",
        )

    dest_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(template_path, dest_path)

    wb = load_workbook(dest_path)
    ws = wb.active
    highlight = PatternFill(fill_type="solid", fgColor="FFFF00")

    # Set partner TADIG.
    ws["C2"] = plan.partner_tadig

    # SMS-MO update (always).
    _write_sms_row(ws, SMS_MO_ROW, rate=plan.sms_mo_rate, currency=plan.currency, is_discount=plan.is_discount, highlight=highlight)

    # SMS-MT update only when provided (> 0).
    if plan.sms_mt_rate and plan.sms_mt_rate > 0:
        _write_sms_row(ws, SMS_MT_ROW, rate=plan.sms_mt_rate, currency=plan.currency, is_discount=plan.is_discount, highlight=highlight)

    wb.save(dest_path)


def _truncate(text: str, limit: int) -> Tuple[str, bool]:
    if len(text) <= limit:
        return text, False
    return text[:limit], True


def _normalize_text(text: str) -> str:
    return (text or "").replace("\x00", "").strip()


def _truncate_inputs(
    agreement_text: str, standards: List[Tuple[str, str]]
) -> Tuple[str, List[Tuple[str, str]], Dict[str, Any]]:
    meta: Dict[str, Any] = {
        "max_chars_per_file": MAX_CHARS_PER_FILE,
        "max_total_chars": MAX_TOTAL_CHARS,
        "truncated": {"agreement": False, "standards": {}},
    }

    agreement_text = _normalize_text(agreement_text)
    agreement_text, truncated_agreement = _truncate(agreement_text, MAX_CHARS_PER_FILE)
    meta["truncated"]["agreement"] = truncated_agreement

    normalized_standards: List[Tuple[str, str]] = []
    for filename, text in standards:
        text = _normalize_text(text)
        text, trunc = _truncate(text, MAX_CHARS_PER_FILE)
        meta["truncated"]["standards"][filename] = trunc
        normalized_standards.append((filename, text))

    total = len(agreement_text) + sum(len(t) for _, t in normalized_standards)
    if total <= MAX_TOTAL_CHARS or not normalized_standards:
        return agreement_text, normalized_standards, meta

    remaining_budget = max(0, MAX_TOTAL_CHARS - len(agreement_text))
    per_standard_budget = max(5_000, remaining_budget // len(normalized_standards))

    final_standards: List[Tuple[str, str]] = []
    for filename, text in normalized_standards:
        text2, trunc2 = _truncate(text, per_standard_budget)
        meta["truncated"]["standards"][filename] = bool(
            meta["truncated"]["standards"][filename] or trunc2
        )
        final_standards.append((filename, text2))

    return agreement_text, final_standards, meta


def _build_user_prompt(
    *, agreement_filename: str, agreement_text: str, standards: List[Tuple[str, str]]
) -> str:
    standards_block = []
    for filename, text in standards:
        standards_block.append(f"--- STANDARD FILE: {filename} ---\n{text}")

    standards_text = "\n\n".join(standards_block)

    return f"""Agreement filename: {agreement_filename}

Agreement text:
{agreement_text}

Standard IOT texts (each labeled by filename):
{standards_text}

Instruction:
Generate loader mapping between agreement clauses and standard templates.
Return JSON matching the required schema exactly.
"""


def _parse_and_validate_loader_json(raw: str) -> LoaderOutput:
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid JSON: {exc}") from exc

    try:
        return LoaderOutput.model_validate(data)
    except Exception as exc:
        raise ValueError(f"JSON did not match schema: {exc}") from exc


async def generate_loader_artifacts(
    *,
    agreement_path: Path,
    standard_paths: List[Path],
    agreement_id: str,
    batch_id: str,
    model: str,
) -> GenerateLoaderResponse:
    started_at = _utc_now()

    try:
        agreement_text = extract_text(agreement_path)
    except TextExtractionError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    standards: List[Tuple[str, str]] = []
    for p in standard_paths:
        try:
            standards.append((p.name, extract_text(p)))
        except TextExtractionError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc

    agreement_text, standards, trunc_meta = _truncate_inputs(agreement_text, standards)

    agreement_filename = agreement_path.name
    user_prompt = _build_user_prompt(
        agreement_filename=agreement_filename, agreement_text=agreement_text, standards=standards
    )

    raw_json, usage = await generate_json_object(
        model=model,
        system_prompt=SYSTEM_PROMPT,
        user_prompt=user_prompt,
        retries=2,
    )

    loader: Optional[LoaderOutput] = None
    parse_error: Optional[str] = None
    try:
        loader = _parse_and_validate_loader_json(raw_json)
    except ValueError as exc:
        parse_error = str(exc)

    if loader is None:
        fix_prompt = (
            "The previous output was invalid JSON or did not match the required schema.\n"
            "Return ONLY the corrected JSON object (no markdown, no commentary).\n\n"
            f"Previous output:\n{raw_json}\n"
        )
        raw_json2, usage2 = await generate_json_object(
            model=model,
            system_prompt=SYSTEM_PROMPT,
            user_prompt=fix_prompt,
            retries=1,
        )
        if usage and usage2:
            usage = {"first_call": usage, "retry_call": usage2}
        elif usage2:
            usage = {"retry_call": usage2}

        try:
            loader = _parse_and_validate_loader_json(raw_json2)
            raw_json = raw_json2
        except ValueError as exc:
            raise HTTPException(
                status_code=502,
                detail=f"Model returned invalid JSON after retry. Last error: {exc}. First error: {parse_error}",
            ) from exc

    if not loader.excel_outputs:
        raise HTTPException(status_code=422, detail="No excel_outputs were returned; cannot generate loader Excel files.")

    agreement_base = safe_agreement_base_name(agreement_path.name, agreement_id)
    output_dir = create_versioned_output_dir(OUTPUT_DIR, agreement_base)

    loader_json_path = output_dir / "loader.json"
    loader_txt_path: Optional[Path] = None
    meta_json_path = output_dir / "meta.json"
    loader_excel_paths: list[Path] = []

    loader_json_path.write_text(
        json.dumps(json.loads(raw_json), indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    # Build Excel outputs using the first provided standard as the template.
    template_source = standard_paths[0] if standard_paths else None
    if template_source is None:
        raise HTTPException(status_code=400, detail="No standard template available to generate Excel output.")

    for plan in loader.excel_outputs:
        dest_filename = plan.filename
        dest_path = output_dir / dest_filename
        _generate_excel_from_template(template_source, dest_path, plan)
        loader_excel_paths.append(dest_path)

    finished_at = _utc_now()
    meta: Dict[str, Any] = {
        "created_at": finished_at,
        "started_at": started_at,
        "agreement_id": agreement_id,
        "batch_id": batch_id,
        "model": model,
        "inputs": {
            "agreement_file": agreement_path.name,
            "standard_files": [p.name for p in standard_paths],
            "truncation": trunc_meta,
            "agreement_chars": len(agreement_text),
            "standards_chars": {fn: len(txt) for fn, txt in standards},
        },
        "openai_usage": usage,
        "outputs": {
            "output_dir": str(output_dir.relative_to(BASE_DIR)),
            "loader_json": str(loader_json_path.relative_to(BASE_DIR)),
            "loader_txt": str(loader_txt_path.relative_to(BASE_DIR)) if loader_txt_path else None,
            "loader_excel": [str(p.relative_to(BASE_DIR)) for p in loader_excel_paths],
            "meta_json": str(meta_json_path.relative_to(BASE_DIR)),
        },
        "excel_plans": [plan.model_dump() for plan in loader.excel_outputs],
    }
    meta_json_path.write_text(json.dumps(meta, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    rel_output_dir = str(output_dir.relative_to(BASE_DIR)) if output_dir.is_absolute() else str(output_dir)
    rel_loader_json = (
        str(loader_json_path.relative_to(BASE_DIR)) if loader_json_path.is_absolute() else str(loader_json_path)
    )
    rel_loader_txt = None
    rel_meta_json = (
        str(meta_json_path.relative_to(BASE_DIR)) if meta_json_path.is_absolute() else str(meta_json_path)
    )
    rel_loader_excel = [
        str(p.relative_to(BASE_DIR)) if p.is_absolute() else str(p) for p in loader_excel_paths
    ]

    summary = {
        "agreement_name": loader.agreement_name,
        "standards_used": loader.standards_used,
        "mappings_count": len(loader.mappings),
        "missing_fields_count": len(loader.missing_fields),
        "excel_files_count": len(loader_excel_paths),
    }

    return GenerateLoaderResponse(
        output_dir=rel_output_dir,
        loader_json_path=rel_loader_json,
        loader_txt_path=rel_loader_txt,
        loader_excel_paths=rel_loader_excel,
        meta_json_path=rel_meta_json,
        summary=summary,
    )
