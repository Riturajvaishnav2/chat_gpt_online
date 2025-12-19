from __future__ import annotations

from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


class TextExtractionError(RuntimeError):
    pass


def extract_text(file_path: Path) -> str:
    suffix = file_path.suffix.lower()
    if suffix == ".txt":
        return _extract_txt(file_path)
    if suffix == ".docx":
        return _extract_docx(file_path)
    if suffix == ".pdf":
        return _extract_pdf(file_path)
    if suffix == ".xlsx":
        return _extract_xlsx(file_path)
    raise TextExtractionError(f"Unsupported file type for text extraction: {file_path.suffix}")


def _extract_txt(file_path: Path) -> str:
    try:
        return file_path.read_text(encoding="utf-8", errors="replace")
    except OSError as exc:
        raise TextExtractionError(f"Failed to read text file: {file_path.name}") from exc


def _extract_docx(file_path: Path) -> str:
    try:
        doc = Document(str(file_path))
    except Exception as exc:  # python-docx raises various exceptions for invalid DOCX
        raise TextExtractionError(f"Failed to read DOCX file: {file_path.name}") from exc

    parts: list[str] = []
    for para in doc.paragraphs:
        txt = (para.text or "").strip()
        if txt:
            parts.append(txt)
    return "\n".join(parts).strip()


def _extract_pdf(file_path: Path) -> str:
    try:
        reader = PdfReader(str(file_path))
    except Exception as exc:
        raise TextExtractionError(f"Failed to read PDF file: {file_path.name}") from exc

    parts: list[str] = []
    for page in reader.pages:
        page_text = page.extract_text() or ""
        page_text = page_text.strip()
        if page_text:
            parts.append(page_text)
    return "\n\n".join(parts).strip()


def _extract_xlsx(file_path: Path) -> str:
    try:
        workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    except Exception as exc:
        raise TextExtractionError(f"Failed to read XLSX file: {file_path.name}") from exc

    parts: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            row_values = [str(cell).strip() for cell in row if cell not in (None, "")]
            if row_values:
                parts.append(" ".join(row_values))

    workbook.close()
    return "\n".join(parts).strip()
